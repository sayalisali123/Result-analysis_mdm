from flask_login import login_manager
import os
import json
import threading
import pandas as pd
import io
from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session, send_file, make_response
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from models import User, ResultAnalysis, MinorDegreeApplication
login_manager = LoginManager()
import firebase_admin
from firebase_admin import credentials, firestore
from datetime import datetime
from dotenv import load_dotenv

# Load environment variables from .env when running locally
load_dotenv()

# Import faculty modules
from extractor import extract_result_data
from mdm_logic import seat_allotment

app = Flask(__name__)
app.config['SECRET_KEY'] = os.getenv('SECRET_KEY', 'dev-secret-key')
# Use /tmp for Vercel, and local uploads/ for Windows/local dev
if os.name == 'nt':
    UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads')
else:
    UPLOAD_FOLDER = '/tmp'

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Initialize Firebase Admin SDK with robust handling
firebase_creds_json = os.environ.get('FIREBASE_CREDENTIALS')
firebase_creds_path = os.environ.get('FIREBASE_CREDENTIALS_PATH')
local_firebase_file = os.path.join(
    os.path.dirname(__file__),
    "student-portal-9f4f6-firebase-adminsdk-fbsvc-614a27b548.json"
)

cred = None

if firebase_creds_json:
    try:
        cred_dict = json.loads(firebase_creds_json)
        cred = credentials.Certificate(cred_dict)
    except Exception as e:
        print(f"Failed to load Firebase credentials from env JSON: {e}")

elif firebase_creds_path and os.path.exists(firebase_creds_path):
    cred = credentials.Certificate(firebase_creds_path)

elif os.path.exists(local_firebase_file):
    cred = credentials.Certificate(local_firebase_file)

else:
    print("Firebase credentials not found. Continuing without Firebase (limited functionality).")

# Initialize Firebase safely
if cred and not firebase_admin._apps:
    firebase_admin.initialize_app(cred)
    db = firestore.client()   # ✅ FIX: removed database_id="native-db"
else:
    # Fallback dummy Firestore to prevent crashes when credentials are missing
    class DummyFirestore:
        def collection(self, *args, **kwargs):
            return self
        def where(self, *args, **kwargs):
            return self
        def stream(self):
            return []
        def add(self, *args, **kwargs):
            pass
        def document(self, *args, **kwargs):
            return self
        def get(self):
            class DummyDoc:
                exists = False
                def to_dict(self):
                    return {}
                id = ''
            return DummyDoc()
        def set(self, *args, **kwargs):
            pass

    db = DummyFirestore()

login_manager.init_app(app)

@login_manager.user_loader
def load_user(user_id):
    try:
        user_ref = db.collection('users').document(str(user_id)).get()
        if user_ref and user_ref.exists:  # type: ignore
            data = user_ref.to_dict()  # type: ignore
            return User.from_dict(data, user_ref.id)  # type: ignore
    except Exception as e:
        print(f"Error loading user: {e}")
    return None

# --- SHARED / PUBLIC ROUTES ---

@app.route('/')
def index():
    return render_template('index.html')

# --- STUDENT ROUTES ---

@app.route('/student_login', methods=['GET', 'POST'])
def student_login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
        
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            flash('Username and password are required.', 'error')
            return render_template('login.html')

        users_ref = db.collection('users').where('username', '==', username).stream()
        user_doc = None
        for doc in users_ref:
            user_doc = doc
            break
        
        if user_doc:
            user = User.from_dict(user_doc.to_dict(), user_doc.id)
            if password and check_password_hash(user.password_hash, password):
                login_user(user)
                return redirect(url_for('dashboard'))
            else:
                flash('Login failed. Check your username and password.', 'error')
        else:
            flash('Login failed. Check your username and password.', 'error')
    
    return render_template('login.html')

@app.route('/student_register', methods=['GET', 'POST'])
def student_register():
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')

        if not username or not password:
            flash('Username and password are required.', 'error')
            return render_template('register.html')

        users_ref = db.collection('users').where('username', '==', username).stream()
        user_exists = any(True for _ in users_ref)
        
        if user_exists:
            flash('Username already exists.', 'error')
            return redirect(url_for('student_register'))

        # Create new user document
        user_data = {
            'username': username,
            'password_hash': generate_password_hash(password) if password else None
        }
        db.collection('users').add(user_data)

        flash('Registration successful. Please log in.', 'success')
        return redirect(url_for('student_login'))

    return render_template('register.html')

@app.route('/logout')
def logout():
    if current_user.is_authenticated:
        logout_user()
    session.pop('faculty_logged_in', None)
    return redirect(url_for('index'))

@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html')

@app.route('/student_results')
@login_required
def student_results():
    # Fetch user's existing records
    result_ref = db.collection('result_analysis').where('user_id', '==', str(current_user.id)).stream()
    result_submissions = [ResultAnalysis.from_dict(doc.to_dict(), doc.id) for doc in result_ref]
    return render_template('student_results.html', results=result_submissions)

@app.route('/student_mdm')
@login_required
def student_mdm():
    # Fetch user's existing records
    mdm_ref = db.collection('mdm_preferences').where('user_id', '==', str(current_user.id)).stream()
    mdm_submissions = [MinorDegreeApplication.from_dict(doc.to_dict(), doc.id) for doc in mdm_ref]
    return render_template('student_mdm.html', mdm=mdm_submissions)

def process_result_background(filepath, doc_id):
    """Background task to extract PDF and update the existing result_analysis document"""
    try:
        data = extract_result_data(pdf_path=filepath)
        if data:
            meta = data['metadata']
            
            # Update the existing document in result_analysis
            doc_ref = db.collection("result_analysis").document(doc_id)
            doc_ref.set({
                "name": meta.get('name', 'UNKNOWN_NAME'),
                "prn": meta.get('prn', 'UNKNOWN_PRN'),
                "sem": meta.get('sem', 'UNKNOWN_SEM'),
                "dept": meta.get('dept', 'UNKNOWN_DEPT'),
                "subjects": data['subjects'],
                "summary": data['summary'],
                "last_updated": datetime.now()
            }, merge=True)
            print(f"Success: Processed result for document {doc_id}")
    except Exception as e:
        print(f"Background Process Error: {e}")

@app.route('/submit_result', methods=['POST'])
@login_required
def submit_result():
    student_class = request.form.get('class')
    roll_no = request.form.get('roll_no')
    department = request.form.get('department')
    
    if 'result_pdf' not in request.files:
        flash('No file part', 'error')
        return redirect(url_for('dashboard'))
    
    file = request.files['result_pdf']
    if not file or file.filename == '':
        flash('No selected file', 'error')
        return redirect(url_for('dashboard'))
        
    if file and file.filename and file.filename.endswith('.pdf'):
        filename = secure_filename(f"{current_user.id}_{roll_no}_{file.filename}")
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Save student submission record in result_analysis
        result_data = {
            'user_id': str(current_user.id),
            'student_class': student_class,
            'roll_no': roll_no,
            'department': department,
            'pdf_filename': filename,
            'last_updated': datetime.now()
        }
        _, doc_ref = db.collection('result_analysis').add(result_data)
        
        # Trigger background extraction, passing the doc_id to update
        process_result_background(filepath, doc_ref.id)

        flash('Result submitted successfully! It is now being analyzed.', 'success')
    else:
        flash('Expected a PDF file.', 'error')
        
    return redirect(url_for('student_results'))

@app.route('/submit_mdm', methods=['POST'])
@login_required
def submit_mdm():
    prn_no = request.form.get('prn_no')
    current_department = request.form.get('current_department')
    total_marks = request.form.get('total_marks', 0)
    preference_1 = request.form.get('preference_1')
    preference_2 = request.form.get('preference_2')
    preference_3 = request.form.get('preference_3')
    preference_4 = request.form.get('preference_4')
    
    # Store in mdm_preferences collection (Unified)
    mdm_data = {
        'user_id': str(current_user.id),
        'Name': 'Student', # Optional: could fetch from user profile
        'PRN': prn_no,
        'Branch': current_department,
        'Total Marks': float(total_marks) if total_marks else 0,
        'PREFERENCE 1': preference_1,
        'PREFERENCE 2': preference_2,
        'PREFERENCE 3': preference_3,
        'PREFERENCE 4': preference_4,
        'submitted_at': datetime.now()
    }
    db.collection('mdm_preferences').add(mdm_data)
    
    flash('Minor Degree Preferences submitted successfully!', 'success')
    return redirect(url_for('student_mdm'))

# --- FACULTY ROUTES ---

from functools import wraps

def faculty_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('faculty_logged_in'):
            return redirect(url_for('faculty_login'))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/faculty_login', methods=['GET'])
def faculty_login():
    return render_template('faculty_login.html')

@app.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    if not data:
        return jsonify({"detail": "Invalid request"}), 400
    
    username = data.get('username')
    password = data.get('password')
    
    if username == "admin" and password == "admin":
        session['faculty_logged_in'] = True
        return jsonify({"status": "success", "message": "Login successful"})
    
    return jsonify({"detail": "Invalid credentials"}), 401

@app.route('/faculty_dashboard')
@faculty_required
def faculty_dashboard():
    return render_template('faculty_dashboard.html')

@app.route('/mdm')
@faculty_required
def faculty_mdm():
    return render_template('faculty_mdm.html')

@app.route('/results_view')
@faculty_required
def faculty_results_view():
    return render_template('faculty_results.html')

@app.route('/api/results')
@faculty_required
def api_results():
    """Fetches all processed results from result_analysis collection."""
    try:
        all_results = []
        # Unified collection is result_analysis
        results_ref = db.collection('result_analysis').stream()
        
        for doc in results_ref:
            doc_data = doc.to_dict()
            if not doc_data.get('summary'): # Skip if not processed yet
                continue
                
            doc_data['id'] = doc.id
            # Ensure keys exist for the frontend table
            doc_data['prn'] = doc_data.get('prn', doc_data.get('roll_no', 'N/A'))
            doc_data['dept'] = doc_data.get('dept', doc_data.get('department', 'N/A'))
            doc_data['sem'] = doc_data.get('sem', doc_data.get('student_class', 'N/A'))
            
            # Fix timestamp serialization
            if 'last_updated' in doc_data and doc_data['last_updated']:
                 doc_data['last_updated'] = str(doc_data['last_updated'])
            
            all_results.append(doc_data)
        
        return jsonify({"status": "success", "data": all_results})
    except Exception as e:
        print(f"Error fetching results: {e}")
        return jsonify({"detail": "Failed to fetch results"}), 500

@app.route('/api/download_results_excel')
@faculty_required
def download_results_excel():
    """Generates and downloads the detailed result analysis as an Excel file."""
    try:
        results_ref = db.collection('result_analysis').stream()
        all_data = []
        unique_sub_cats = set() # (Subject Name, Category)
        
        for doc in results_ref:
            d = doc.to_dict()
            if not d.get('summary'): continue
            all_data.append(d)
            for sub in d.get('subjects', []):
                unique_sub_cats.add((sub['name'], sub['category']))
                
        if not all_data:
            return "No data to download", 404
            
        sorted_sub_cats = sorted(list(unique_sub_cats))
        
        # Build helper structures for Excel header merging
        subjectsMap = {}
        for sub_name, cat in sorted_sub_cats:
            if sub_name not in subjectsMap:
                subjectsMap[sub_name] = []
            subjectsMap[sub_name].append(cat)
        sorted_subjects = sorted(list(subjectsMap.keys()))
        
        # Prepare Data for DataFrame
        rows = []
        for d in all_data:
            row = {
                ('Student', 'PRN'): d.get('prn', 'N/A'),
                ('Student', 'Name'): d.get('name', 'N/A')
            }
            # Add subject data
            for sub_name, cat in sorted_sub_cats:
                # Find matching subject entry
                match = next((s for s in d.get('subjects', []) if s['name'] == sub_name and s['category'] == cat), None)
                row[(sub_name, f"{cat} Score")] = match['score'] if match else '-'
                row[(sub_name, f"{cat} Status")] = match['status'] if match else '-'
            
            # Add overall
            row[('OVERALL', 'Total')] = f"{d['summary'].get('total_obtained', 0)} / {d['summary'].get('out_of', 0)}"
            row[('OVERALL', 'Percentage')] = f"{d['summary'].get('percentage', 0)}%"
            rows.append(row)
            
        # Create Excel in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # We write without headers first, starting from row 3 (0-indexed 2)
            # First, prepare a flat dataframe for data only
            flat_rows = []
            for d in all_data:
                row_data = [d.get('prn', 'N/A'), d.get('name', 'N/A')]
                for sub_name, cat in sorted_sub_cats:
                    match = next((s for s in d.get('subjects', []) if s['name'] == sub_name and s['category'] == cat), None)
                    row_data.append(match['score'] if match else '-')
                    row_data.append(match['status'] if match else '-')
                row_data.append(f"{d['summary'].get('total_obtained', 0)} / {d['summary'].get('out_of', 0)}")
                row_data.append(f"{d['summary'].get('percentage', 0)}%")
                flat_rows.append(row_data)
            
            df_data = pd.DataFrame(flat_rows)
            df_data.to_excel(writer, index=False, header=False, startrow=2, sheet_name='Sheet1')
            
            # Apply formatting and headers using openpyxl
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            
            from openpyxl.styles import Font, Alignment, PatternFill
            header_fill = PatternFill(start_color="3F51B5", end_color="3F51B5", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            center_align = Alignment(horizontal="center", vertical="center")
            
            # Write Row 1 (Subject Names)
            worksheet.cell(row=1, column=1, value="Student Info")
            worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            
            col_ptr = 3
            for sub_name, cat_count in [(s, len(list(subjectsMap[s]))) for s in sorted_subjects]:
                worksheet.cell(row=1, column=col_ptr, value=sub_name)
                if cat_count > 1 or True: # Always merge for Score/Status pair
                    worksheet.merge_cells(start_row=1, start_column=col_ptr, end_row=1, end_column=col_ptr + (cat_count * 2) - 1)
                col_ptr += (cat_count * 2)
            
            worksheet.cell(row=1, column=col_ptr, value="OVERALL")
            worksheet.merge_cells(start_row=1, start_column=col_ptr, end_row=1, end_column=col_ptr + 1)
            
            # Write Row 2 (Components)
            worksheet.cell(row=2, column=1, value="PRN")
            worksheet.cell(row=2, column=2, value="Name")
            
            col_ptr = 3
            for sub_name in sorted_subjects:
                cats = sorted(list(subjectsMap[sub_name]))
                for cat in cats:
                    worksheet.cell(row=2, column=col_ptr, value=f"{cat} Score")
                    worksheet.cell(row=2, column=col_ptr + 1, value="Status")
                    col_ptr += 2
            
            worksheet.cell(row=2, column=col_ptr, value="Total")
            worksheet.cell(row=2, column=col_ptr + 1, value="Percentage")
            
            # Style headers
            for r in [1, 2]:
                for c in range(1, col_ptr + 2):
                    cell = worksheet.cell(row=r, column=c)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align
                    if r == 2:
                        worksheet.column_dimensions[cell.column_letter].width = 15

        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'Result_Analysis_Report_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
    except Exception as e:
        print(f"Error downloading Results Excel: {e}")
        return str(e), 500

@app.route('/api/mdm_preferences')
@faculty_required
def api_mdm_preferences():
    """Fetches MDM preferences and merges with academic results."""
    try:
        # Fetch all results to map PRN to Marks/Percentage
        results_ref = db.collection('result_analysis').stream()
        results_map = {}
        for r_doc in results_ref:
            r_data = r_doc.to_dict()
            prn = r_data.get('prn')
            if prn and r_data.get('summary'):
                results_map[str(prn)] = r_data['summary']

        mdm_ref = db.collection('mdm_preferences').stream()
        mdm_data = []
        for doc in mdm_ref:
            data = doc.to_dict()
            if data:
                data['id'] = doc.id
                prn = str(data.get('PRN', ''))
                
                # Merge marks and percentage from results if available
                if prn in results_map:
                    data['Marks'] = results_map[prn].get('total_obtained', 0)
                    data['Percentage'] = f"{results_map[prn].get('percentage', 0)}%"
                else:
                    data['Marks'] = data.get('Total Marks', 0)
                    data['Percentage'] = 'N/A'
                
                mdm_data.append(data)
            
        if not mdm_data:
            return jsonify({"status": "success", "data": []})
            
        df = pd.DataFrame(mdm_data)
        # Rename Branch to Current Dept for clarity in logic (it will be renamed back if needed)
        if 'Branch' in df.columns:
            df = df.rename(columns={'Branch': 'Current Dept'})
            
        allocated_df = seat_allotment(df)
        allocated_data = allocated_df.to_dict(orient='records')
        
        return jsonify({"status": "success", "data": allocated_data})
    except Exception as e:
        print(f"Error fetching MDM preferences: {e}")
        return jsonify({"detail": "Failed to fetch MDM preferences"}), 500

@app.route('/api/download_mdm_excel')
@faculty_required
def download_mdm_excel():
    """Generates and downloads the MDM allocation results as an Excel file."""
    try:
        # Fetch data (reusing the logic from api_mdm_preferences)
        results_ref = db.collection('result_analysis').stream()
        results_map = {str(r.to_dict().get('prn')): r.to_dict().get('summary') for r in results_ref if r.to_dict().get('summary')}
        
        mdm_ref = db.collection('mdm_preferences').stream()
        mdm_data = []
        for doc in mdm_ref:
            data = doc.to_dict()
            if data:
                prn = str(data.get('PRN', ''))
                if prn in results_map:
                    data['Marks'] = results_map[prn].get('total_obtained', 0)
                    data['Percentage'] = f"{results_map[prn].get('percentage', 0)}%"
                else:
                    data['Marks'] = data.get('Total Marks', 0)
                    data['Percentage'] = 'N/A'
                mdm_data.append(data)
        
        if not mdm_data:
            return "No data to download", 404
            
        df = pd.DataFrame(mdm_data)
        if 'Branch' in df.columns:
            df = df.rename(columns={'Branch': 'Current Dept'})
            
        allocated_df = seat_allotment(df)
        
        # Select and order columns for the Excel file as per requirements
        export_cols = ['PRN', 'Name', 'Marks', 'Percentage', 'Current Dept', 'Allocated Minor', 'Preference Used', 'Status']
        # Filter only existing columns to avoid errors
        export_cols = [c for c in export_cols if c in allocated_df.columns]
        export_df = allocated_df[export_cols]
        
        # Create Excel in memory
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            export_df.to_excel(writer, index=False, sheet_name='MDM Allocation')
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'MDM_Allocation_{datetime.now().strftime("%Y%m%d_%H%M")}.xlsx'
        )
    except Exception as e:
        print(f"Error downloading Excel: {e}")
        return str(e), 500

@app.route('/analyze', methods=['POST'])
@faculty_required
def analyze_endpoint():
    """Fallback endpoint for manual analysis"""
    return jsonify({"status": "Processing", "message": "Endpoint preserved, but handled automatically on student upload."})

if __name__ == '__main__':
    app.run(debug=True, port=5001)
